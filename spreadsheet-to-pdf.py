"""
spreadsheet-to-pdf.py
Conversione di file foglio elettronico (Excel e OpenDocument) multifoglio
in PDF, con best-fit delle colonne e gestione robusta del caso "grande
numero di colonne".

Formati di input supportati:
  - .xlsx / .xlsm (Microsoft Excel)
  - .ods / .fods  (OpenDocument Spreadsheet)

Strategia:
  1. (Solo per .ods/.fods) LibreOffice headless converte preventivamente
     in .xlsx, perché openpyxl gestisce nativamente solo i formati Excel.
  2. openpyxl: per ogni foglio calcola il best-fit delle colonne in base al
     contenuto effettivo e imposta i parametri di stampa nel file XLSX
     (orientamento, formato carta, fit-to-width, margini, ripetizione
     intestazioni).
  3. LibreOffice in modalità headless (`soffice --convert-to pdf`):
     converte l'XLSX preparato in PDF rispettando le impostazioni di stampa
     embedded.

Dipendenze (verificate attivamente mantenute):
  - openpyxl >= 3.1.5  (https://openpyxl.readthedocs.io/)
  - LibreOffice >= 7.x con `soffice` in PATH o percorso fornito da CLI.

Installazione:
    pip install openpyxl
    # LibreOffice: https://www.libreoffice.org/download/

Uso base:
    python spreadsheet-to-pdf.py file.xlsx
    python spreadsheet-to-pdf.py file.ods            # supporto trasparente OpenDocument
    python spreadsheet-to-pdf.py file.xlsx -o output.pdf
    python spreadsheet-to-pdf.py file.xlsx --paper a3 --orientation landscape
    python spreadsheet-to-pdf.py *.xlsx *.ods --outdir ./pdf/

Esempi:
    # Conversione singola, formato automatico (cap A3 + wrap se serve):
    python spreadsheet-to-pdf.py report_annuale.xlsx --repeat-cols 3

    # Forza A3 orizzontale, larghezza max colonna 60 caratteri:
    python spreadsheet-to-pdf.py dati.ods --paper a3 --max-col-width 60

    # Registro largo a video: cap A3 con wrap automatico delle colonne
    # in eccesso, mantenendo ID/Cognome/Nome ripetuti per ogni blocco:
    python spreadsheet-to-pdf.py registro.xlsx --repeat-cols 3

    # Wrap esplicito (forza il riassetto anche se la scala basterebbe):
    python spreadsheet-to-pdf.py registro.xlsx --wide-strategy wrap-rows --repeat-cols 3

    # Allargo il cap a A2 se ho un plotter o un monitor grande:
    python spreadsheet-to-pdf.py registro.xlsx --max-paper a2 --repeat-cols 3

    # Conversione batch con debug (mantiene file temporaneo):
    python spreadsheet-to-pdf.py *.xlsx *.ods --outdir ./pdf --keep-temp -v

Gestione fogli larghi (--wide-strategy, default "auto"):
    auto         se sta in --max-paper a scala >= --min-scale (95%), comprime;
                 altrimenti wrappa le colonne in righe sottostanti
    scale-down   comprime tutto su 1 pagina (testo piccolo)
    multipage    font pieno, colonne in eccesso su pagine ORIZZONTALI successive
    fixed-scale  scala fissa --fixed-scale %, impaginazione naturale
    wrap-rows    ristruttura il foglio: colonne in eccesso vanno in righe
                 sottostanti, con le prime --repeat-cols colonne ripetute
                 in ogni blocco (un blocco per pagina, separato da page break)
"""

import argparse
import os
import platform
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("❌ openpyxl non installato. Esegui: pip install openpyxl")


# Codici openpyxl per i formati carta più usati
# Riferimento: https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.page.html
PAPER_SIZES = {
    "letter":  1,   # 8.5 x 11 in
    "legal":   5,   # 8.5 x 14 in
    "tabloid": 3,   # 11 x 17 in
    "a3":      8,   # 297 x 420 mm
    "a4":      9,   # 210 x 297 mm
    "a5":     11,
    "a2":     66,   # 420 x 594 mm
    "a1":     65,   # 594 x 841 mm
}

# Dimensioni carta in mm (lato_corto, lato_lungo), formato verticale.
# Servono alla strategia "auto" per stimare la scala necessaria al fit.
PAPER_DIMENSIONS_MM = {
    "a5":      (148, 210),
    "a4":      (210, 297),
    "a3":      (297, 420),
    "a2":      (420, 594),
    "a1":      (594, 841),
    "letter":  (216, 279),
    "legal":   (216, 356),
    "tabloid": (279, 432),
}

# Margine orizzontale totale (sinistra + destra) in mm per ciascun preset.
HORIZONTAL_MARGIN_MM = {
    "narrow": 12.7,   # 0.25" + 0.25"
    "normal": 35.6,   # 0.7"  + 0.7"
    "wide":   50.8,   # 1.0"  + 1.0"
}

# Larghezza colonna in unità Excel = (caratteri + padding) * factor.
# Excel usa la larghezza del carattere "0" del font default come unità.
DEFAULT_FONT_FACTOR = 1.10  # margine extra per font non monospace
DEFAULT_PADDING = 2
DEFAULT_MIN_COL_WIDTH = 8.0
DEFAULT_MAX_COL_WIDTH = 50.0

# Estensioni gestite nativamente da openpyxl (preprocessing diretto).
NATIVE_EXTENSIONS = {".xlsx", ".xlsm"}
# Estensioni che richiedono conversione preliminare a .xlsx via LibreOffice.
OPENDOCUMENT_EXTENSIONS = {".ods", ".fods"}
SUPPORTED_EXTENSIONS = NATIVE_EXTENSIONS | OPENDOCUMENT_EXTENSIONS


# ─────────────────────────────────────────────────────────────────────────────
# Rilevamento LibreOffice
# ─────────────────────────────────────────────────────────────────────────────

def find_soffice(user_path: str | None = None) -> str:
    """Trova l'eseguibile di LibreOffice (soffice).

    Cerca nell'ordine:
      1. Percorso fornito dall'utente.
      2. `soffice` nel PATH.
      3. Percorsi tipici per OS.
    """
    if user_path:
        if Path(user_path).is_file():
            return user_path
        sys.exit(f"❌ Percorso soffice non valido: {user_path}")

    # PATH
    for name in ("soffice", "libreoffice"):
        found = shutil.which(name)
        if found:
            return found

    # Percorsi tipici
    system = platform.system()
    candidates: list[str] = []
    if system == "Windows":
        candidates = [
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        ]
    elif system == "Darwin":  # macOS
        candidates = [
            "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        ]
    else:  # Linux/BSD
        candidates = [
            "/usr/bin/soffice",
            "/usr/bin/libreoffice",
            "/snap/bin/libreoffice",
            "/opt/libreoffice/program/soffice",
        ]

    for c in candidates:
        if Path(c).is_file():
            return c

    sys.exit(
        "❌ LibreOffice non trovato.\n"
        "   Installa da https://www.libreoffice.org/download/\n"
        "   oppure usa --soffice per specificare il percorso."
    )


# ─────────────────────────────────────────────────────────────────────────────
# Best-fit colonne
# ─────────────────────────────────────────────────────────────────────────────

def compute_text_width(value, number_format: str | None = None) -> int:
    """Stima la larghezza in caratteri del valore di una cella.

    Gestisce:
      - Valori None → 0
      - Testi multilinea → lunghezza massima per riga
      - Numeri/date → usa il formato di visualizzazione se disponibile
    """
    if value is None:
        return 0

    # Per date/datetime usa una stima ragionevole se non c'è formato
    if hasattr(value, "isoformat"):
        return max(10, len(str(value)))

    text = str(value)
    if "\n" in text:
        return max(len(line) for line in text.split("\n"))
    return len(text)


def compute_column_widths(
    ws,
    max_width: float = DEFAULT_MAX_COL_WIDTH,
    min_width: float = DEFAULT_MIN_COL_WIDTH,
    padding: int = DEFAULT_PADDING,
    factor: float = DEFAULT_FONT_FACTOR,
    sample_rows: int | None = None,
) -> dict[int, float]:
    """Calcola e applica al worksheet la larghezza best-fit per ogni colonna.

    Args:
        ws: worksheet di openpyxl.
        max_width: clamp superiore (evita colonne enormi per testi lunghi).
        min_width: clamp inferiore (per intestazioni o colonne sparse).
        padding: caratteri extra di respiro.
        factor: fattore correttivo per font non monospace.
        sample_rows: se non None, campiona solo le prime N righe (utile per
                     fogli enormi).

    Returns:
        dict {indice_colonna_1based: larghezza_calcolata}
    """
    widths: dict[int, int] = {}
    max_row = ws.max_row
    if sample_rows is not None:
        max_row = min(max_row, sample_rows)

    # Scansione efficiente per colonna usando iter_cols
    for col_idx, col in enumerate(
        ws.iter_cols(min_row=1, max_row=max_row,
                     min_col=1, max_col=ws.max_column,
                     values_only=False),
        start=1,
    ):
        max_len = 0
        for cell in col:
            w = compute_text_width(cell.value, cell.number_format)
            if w > max_len:
                max_len = w
        widths[col_idx] = max_len

    # Applica le larghezze al worksheet
    final = {}
    for col_idx, length in widths.items():
        width = (length + padding) * factor
        width = max(min_width, min(width, max_width))
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = width
        final[col_idx] = width
    return final


# ─────────────────────────────────────────────────────────────────────────────
# Impostazioni di stampa
# ─────────────────────────────────────────────────────────────────────────────

def auto_select_paper(num_cols: int, total_width: float, max_paper: str = "a3") -> tuple[str, str]:
    """Seleziona formato carta e orientamento, rispettando un cap massimo
    (default A3, ottimale per consultazione a video).

    `total_width` è in unità-carattere Excel, non in mm; serve come
    correttivo per fogli con colonne particolarmente larghe.
    """
    avg = total_width / max(num_cols, 1)
    weight = num_cols * (1 + max(0, (avg - 12) / 20))

    paper_ladder = [("a4", 8), ("a3", 20), ("a2", 40), ("a1", float("inf"))]
    paper_order = [p for p, _ in paper_ladder]

    if max_paper not in paper_order:
        max_paper = paper_order[-1]
    cap_idx = paper_order.index(max_paper)

    for paper, threshold in paper_ladder[: cap_idx + 1]:
        if weight <= threshold:
            return paper, "landscape"
    # Oltre la soglia del cap: restituiamo il cap stesso. La strategia
    # wrap-rows (o scale-down) deciderà cosa fare con l'overflow.
    return max_paper, "landscape"


def printable_width_mm(paper: str, orientation: str, margins: str) -> float:
    """Larghezza stampabile (mm) per formato/orientamento/margini dati."""
    short, long = PAPER_DIMENSIONS_MM.get(paper, PAPER_DIMENSIONS_MM["a4"])
    width = long if orientation == "landscape" else short
    return width - HORIZONTAL_MARGIN_MM.get(margins, 12.7)


def char_to_mm(char_width: float) -> float:
    """Conversione approssimativa unità-carattere Excel → mm.
    Stessa euristica di estimate_fit_scale (Calibri 11 @ 96 DPI:
    ~7 px per unità + ~5 px di padding). Serve a wrap_sheet_into_rows
    per stimare quante colonne stanno in un blocco.
    """
    return (char_width * 7 + 5) / 96 * 25.4


def estimate_fit_scale(
    total_char_width: float,
    num_cols: int,
    paper: str,
    orientation: str,
    margins: str,
) -> int:
    """Stima la scala % che `fitToWidth=1` produrrebbe per far stare tutte
    le colonne nella larghezza stampabile.

    La conversione unità-carattere → mm è approssimativa (basata su
    Calibri 11 @ 96 DPI: ~7 px per unità di larghezza + ~5 px di padding
    per colonna). È sufficiente per la SOGLIA decisionale della strategia
    "auto", non per un layout preciso.

    Returns:
        Scala percentuale stimata (100 = nessuna riduzione necessaria).
    """
    total_px = total_char_width * 7 + num_cols * 5
    total_mm = total_px / 96 * 25.4
    printable = printable_width_mm(paper, orientation, margins)
    if total_mm <= printable:
        return 100
    return max(1, int(printable / total_mm * 100))


def wrap_sheet_into_rows(
    ws,
    anchor_cols: int,
    available_data_width_mm: float,
    col_widths_chars: dict[int, float],
    safety_margin: float = 0.12,
    page_height_mm: float | None = None,
    separator_rows: int = 1,
    separator_height_pt: float = 8.0,
    safety_v: float = 0.05,
    min_row_height_pt: float = 12.0,
    block_titles: list[str] | None = None,
) -> tuple[list[int], int]:
    """Sposta le colonne in eccesso nelle righe successive dello stesso foglio.

    Da un foglio largo:

        H_a1 H_a2 H_a3  H_d1 H_d2 H_d3 ... H_dM
        ...                                       (R righe)

    produce un foglio "wrappato" verticalmente:

        H_a1 H_a2 H_a3  H_d1 H_d2 H_d3
        ...                                       (R righe del blocco 1)
        --- page break ---
        H_a1 H_a2 H_a3  H_d4 H_d5 H_d6
        ...                                       (R righe del blocco 2)
        --- page break ---
        ...

    Le prime `anchor_cols` colonne (es. ID/Cognome/Nome) vengono ripetute
    in ciascun blocco così le righe restano identificabili. Tra blocchi
    consecutivi si inserisce un'interruzione di pagina hardware perché
    LibreOffice generi una pagina per blocco.

    Args:
        ws: worksheet openpyxl da modificare in-place.
        anchor_cols: numero di colonne iniziali da ripetere in ogni blocco.
        available_data_width_mm: larghezza disponibile per le colonne-dati
            in un blocco (= larghezza stampabile - larghezza colonne ancora).
        col_widths_chars: larghezza colonne in unità-carattere (output di
            compute_column_widths).
        safety_margin: frazione di larghezza riservata a coprire l'errore
            della stima char_to_mm (default 0.12 = 12%). Senza questo
            margine, blocchi calcolati come "appena entranti" finiscono
            su 2 pagine orizzontali nel PDF reale per via di metriche
            font/padding/bordi che la mia euristica non modella.
        page_height_mm: altezza utile della pagina (mm). Se fornita, vengono
            impilati più blocchi sulla stessa pagina con separatori, mettendo
            page break SOLO dove il blocco successivo non entrerebbe. Senza
            questo parametro si torna al comportamento "un blocco per pagina".
        separator_rows: numero di righe vuote tra blocchi consecutivi nella
            stessa pagina (default 1).
        separator_height_pt: altezza in punti delle righe-separatore (default
            8 pt vs. ~15 pt standard). Compatta lo spazio bianco fra blocchi.
        safety_v: frazione di altezza riservata per sicurezza verticale
            (default 0.05 = 5%). Senza questo margine, blocchi calcolati come
            "appena entranti" finirebbero spezzati a metà fra due pagine
            perché LibreOffice rende le righe leggermente più alte della
            mia stima (font reali, padding, bordi).
        min_row_height_pt: altezza minima accettabile in punti per le righe
            del blocco (default 12). Quando per fare entrare N blocchi per
            pagina servirebbero righe più basse di questa soglia, il packing
            scende a N-1 invece di sacrificare leggibilità.
        block_titles: lista di titoli da aggiungere come riga sopra ciascun
            blocco. I titoli vengono assegnati ai blocchi auto-calcolati nello
            stesso ordine in cui appaiono nella lista. Se sono più dei blocchi
            gli extra vengono ignorati con un warning; se sono meno, i blocchi
            in eccesso non avranno riga-titolo. La composizione dei blocchi
            resta determinata dal greedy fill: questa opzione aggiunge solo
            l'etichetta.

    Returns:
        Tupla `(block_start_rows, num_blocks)`:
          - block_start_rows: righe (1-indexed) dove serve un page break,
            cioè i blocchi che non entrano nella pagina del precedente.
            Lista vuota se tutti i blocchi entrano in 1 pagina.
          - num_blocks: numero totale di blocchi creati dal wrap (utile
            per i messaggi diagnostici).
        Se non c'è stato bisogno di wrap (1 solo blocco basta), ritorna
        `([], 0)`.

    Limitazioni note:
      - Le formule non vengono riadattate (i riferimenti restano alle celle
        originali); pensato per fogli di dati e voti, non per modelli di calcolo.
      - Le celle unite e le formattazioni per-cella nei dati vengono perse;
        si preservano font/fill/allineamento/bordi solo della riga intestazione.
    """
    from openpyxl.worksheet.dimensions import RowDimension

    n_rows = ws.max_row
    n_cols = ws.max_column

    if anchor_cols < 0:
        anchor_cols = 0
    if anchor_cols >= n_cols:
        return [], 0  # tutte le colonne sono "ancora", non c'è nulla da wrappare

    # 1. Leggi tutti i valori in memoria (data[r-1][c-1] = ws.cell(r,c).value)
    data = [
        [ws.cell(r, c).value for c in range(1, n_cols + 1)]
        for r in range(1, n_rows + 1)
    ]

    # 2. Leggi lo stile della riga di intestazione (per riapplicarlo a ogni blocco)
    # NB: openpyxl espone cell.font/fill/... come StyleProxy non hashable;
    # serve una copia esplicita per poter riassegnare gli stili a celle nuove.
    from copy import copy as _copy
    header_styles = [
        (
            _copy(ws.cell(1, c).font),
            _copy(ws.cell(1, c).fill),
            _copy(ws.cell(1, c).alignment),
            _copy(ws.cell(1, c).border),
        )
        for c in range(1, n_cols + 1)
    ]

    # 3. Composizione dei blocchi: greedy fill sulle larghezze in mm.
    # Riduco la larghezza utile col safety_margin per coprire l'errore
    # della stima char_to_mm (font reali, padding, bordi).
    target_width = available_data_width_mm * (1 - safety_margin)
    widths_mm = {c: char_to_mm(col_widths_chars.get(c, 8))
                 for c in range(1, n_cols + 1)}

    blocks: list[list[int]] = []
    current: list[int] = []
    current_w = 0.0
    for col_idx in range(anchor_cols + 1, n_cols + 1):
        w = widths_mm[col_idx]
        if current_w + w > target_width and current:
            blocks.append(current)
            current = [col_idx]
            current_w = w
        else:
            current.append(col_idx)
            current_w += w
    if current:
        blocks.append(current)

    # 3b. Assegna titoli ai blocchi (allineati per indice).
    if block_titles:
        if len(block_titles) > len(blocks):
            print(f"   ⚠️  --block: {len(block_titles)} titoli forniti per "
                  f"{len(blocks)} blocchi auto-calcolati; "
                  f"extra ignorati: {block_titles[len(blocks):]}")
        elif len(block_titles) < len(blocks):
            print(f"   ⚠️  --block: {len(block_titles)} titoli forniti per "
                  f"{len(blocks)} blocchi: gli ultimi {len(blocks) - len(block_titles)} "
                  f"blocchi non avranno titolo")
        assigned_titles: list[str | None] = [
            block_titles[i] if i < len(block_titles) else None
            for i in range(len(blocks))
        ]
    else:
        assigned_titles = [None] * len(blocks)

    if len(blocks) <= 1 and not block_titles:
        return [], 0   # tutto sta in un blocco, wrap non necessario
    # NB: se block_titles è fornito procediamo anche con 1 solo blocco, così
    # la riga-titolo colorata viene comunque scritta sopra il blocco unico
    # (caso tipico: il foglio entra in una pagina ma l'utente vuole comunque
    # un titolo identificativo, es. "Pagella finale" sopra la tabella).

    # 4. Stima altezza media righe (deve essere fatta PRIMA di delete_rows;
    # default openpyxl quando non specificata: 15 pt).
    heights = []
    for r in range(1, n_rows + 1):
        rd = ws.row_dimensions.get(r)
        if rd is not None and rd.height is not None:
            heights.append(rd.height)
    row_height_pt = sum(heights) / len(heights) if heights else 15.0

    # 5. Pulizia del foglio (dati e dimensioni riga)
    ws.delete_rows(1, n_rows)
    # delete_rows non garantisce la pulizia delle row_dimensions; lo facciamo
    # esplicitamente perché vogliamo riassegnare altezze ai separatori.
    ws.row_dimensions = type(ws.row_dimensions)(worksheet=ws)

    # 6. Calcolo packing: massimizza blocks_per_page comprimendo l'altezza
    # riga quanto serve, ma mai sotto min_row_height_pt. Applica safety_v
    # come margine contro overflow di rendering.
    has_titles = any(t for t in assigned_titles)
    title_rows = 1 if has_titles else 0
    # Riga titolo (se presente) + intestazione colonne + dati
    block_total_rows = n_rows + title_rows
    default_block_pt = block_total_rows * row_height_pt

    blocks_per_page = 1
    actual_row_height_pt = row_height_pt

    if page_height_mm is not None:
        page_height_pt = page_height_mm / 25.4 * 72
        safe_page_pt = page_height_pt * (1 - safety_v)

        if default_block_pt > safe_page_pt:
            # Un blocco non entra nemmeno da solo: 1 per pagina, mantengo
            # l'altezza di default; LibreOffice spezzerà inevitabilmente.
            # (Per evitarlo, --max-paper più grande o classe più piccola.)
            pass
        else:
            # Cerco il massimo numero di blocchi per pagina:
            # - prima preferisco default_row_height (niente compressione)
            # - poi accetto compressione fino a min_row_height_pt
            for target in range(2, 100):  # cap pratico
                # Altezza riga massima che permetterebbe `target` blocchi
                # nel safe_page_pt, considerando i (target-1) separatori
                max_h = (
                    safe_page_pt - (target - 1) * separator_height_pt
                ) / (target * block_total_rows)

                if max_h >= row_height_pt:
                    # Niente compressione necessaria
                    blocks_per_page = target
                    actual_row_height_pt = row_height_pt
                elif max_h >= min_row_height_pt:
                    # Comprimi al massimo accettabile per questo target
                    blocks_per_page = target
                    actual_row_height_pt = max_h
                else:
                    # Per `target` servirebbe altezza < min_row_height_pt:
                    # fermo qui, mantengo il packing precedente
                    break

    # 7. Scrittura dei blocchi compattati
    from openpyxl.styles import Font, PatternFill, Alignment

    # 7. Espansione blocks → sub_blocks_list: se un blocco è troppo grande
    # per stare in una pagina (es. classe molto numerosa), lo divido in più
    # sub-blocchi consecutivi, ognuno preceduto dal titolo e dall'intestazione
    # delle colonne. Così a inizio di ogni nuova pagina l'utente ha sempre
    # il contesto del blocco visibile, anche quando il blocco straborda.
    data_rows = n_rows - 1   # righe-dati (escludo l'header originale in riga 1)
    title_overhead = 1 if has_titles else 0
    overhead_rows = title_overhead + 1  # titolo (se presente) + header colonne

    if page_height_mm is not None:
        safe_page_pt_eff = page_height_mm / 25.4 * 72 * (1 - safety_v)
        rows_per_page_max = max(1, int(safe_page_pt_eff / actual_row_height_pt))
    else:
        rows_per_page_max = data_rows + overhead_rows  # nessun limite

    if data_rows + overhead_rows > rows_per_page_max:
        # Blocco singolo eccede la pagina: suddivido. Forzo blocks_per_page=1
        # perché ogni sub-blocco occupa già una pagina intera.
        data_rows_per_sub = max(1, rows_per_page_max - overhead_rows)
        blocks_per_page = 1
    else:
        data_rows_per_sub = data_rows

    sub_blocks_list: list[tuple[list[int], str | None, range]] = []
    for block_idx, block_cols in enumerate(blocks):
        title = assigned_titles[block_idx]
        if data_rows <= data_rows_per_sub:
            sub_blocks_list.append((block_cols, title, range(2, n_rows + 1)))
        else:
            n_subs = (data_rows + data_rows_per_sub - 1) // data_rows_per_sub
            for s in range(n_subs):
                start = 2 + s * data_rows_per_sub
                end = min(start + data_rows_per_sub, n_rows + 1)
                if title:
                    sub_title = title if s == 0 else f"{title} (continua)"
                else:
                    sub_title = None
                sub_blocks_list.append((block_cols, sub_title, range(start, end)))

    # 8. Scrittura dei (sub-)blocchi compattati
    anchor_indices = list(range(1, anchor_cols + 1))
    block_start_rows: list[int] = []   # righe dove serve un page break
    out_row = 1
    blocks_on_current_page = 0

    for sub_idx, (block_cols, sub_title, data_range) in enumerate(sub_blocks_list):
        if sub_idx > 0:
            if blocks_on_current_page >= blocks_per_page:
                # Il (sub-)blocco successivo non entra: page break
                block_start_rows.append(out_row)
                blocks_on_current_page = 0
            else:
                # Separatore: righe vuote strette
                for k in range(separator_rows):
                    ws.row_dimensions[out_row + k] = RowDimension(
                        ws, ht=separator_height_pt, customHeight=True,
                    )
                out_row += separator_rows

        src_cols = anchor_indices + block_cols
        total_block_cols = len(src_cols)

        # Riga TITOLO blocco (se applicabile): fascia colorata che attraversa
        # tutte le colonne del blocco. Aiuta a identificare il blocco quando
        # se ne impacchettano più di uno per pagina o quando un blocco si
        # estende su più pagine.
        if sub_title:
            # Spazio a inizio testo come padding visivo dal bordo cella.
            title_cell = ws.cell(out_row, 1, value=f"  {sub_title}")
            title_cell.font = Font(size=12, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill("solid", fgColor="305496")
            title_cell.alignment = Alignment(
                horizontal="left", vertical="center",
            )
            # Merge attraverso tutte le colonne del blocco
            if total_block_cols > 1:
                ws.merge_cells(
                    start_row=out_row, start_column=1,
                    end_row=out_row, end_column=total_block_cols,
                )
            ws.row_dimensions[out_row] = RowDimension(
                ws, ht=actual_row_height_pt, customHeight=True,
            )
            out_row += 1

        # Riga di intestazione delle colonne (stile copiato dall'originale)
        for new_c, src_c in enumerate(src_cols, start=1):
            cell = ws.cell(out_row, new_c, value=data[0][src_c - 1])
            font, fill, alignment, border = header_styles[src_c - 1]
            cell.font = font
            cell.fill = fill
            cell.alignment = alignment
            cell.border = border
        ws.row_dimensions[out_row] = RowDimension(
            ws, ht=actual_row_height_pt, customHeight=True,
        )
        out_row += 1

        # Righe dati di questo sub-blocco
        for src_r in data_range:
            for new_c, src_c in enumerate(src_cols, start=1):
                ws.cell(out_row, new_c, value=data[src_r - 1][src_c - 1])
            ws.row_dimensions[out_row] = RowDimension(
                ws, ht=actual_row_height_pt, customHeight=True,
            )
            out_row += 1

        blocks_on_current_page += 1

    # 9. Larghezze colonne nel nuovo layout
    max_block_size = max(len(b) for b in blocks)

    # Anchor: stessa larghezza dell'originale
    for c in range(1, anchor_cols + 1):
        if c in col_widths_chars:
            ws.column_dimensions[get_column_letter(c)].width = col_widths_chars[c]

    # Posizioni dati: larghezza = max delle colonne-sorgente che vi finiscono
    for pos in range(max_block_size):
        max_w = 0.0
        for block in blocks:
            if pos < len(block):
                w = col_widths_chars.get(block[pos], 8)
                if w > max_w:
                    max_w = w
        new_col_idx = anchor_cols + pos + 1
        ws.column_dimensions[get_column_letter(new_col_idx)].width = max_w

    return block_start_rows, len(blocks)


def prepend_header_row(ws, title: str, row_height_pt: float = 18.0) -> None:
    """Antepone una riga-titolo colorata in cima al foglio, spostando giù
    tutti i dati esistenti.

    A differenza dei titoli-blocco di wrap_sheet_into_rows (uno per blocco),
    questa è una singola riga-titolo globale del foglio: utile con strategie
    diverse da wrap-rows (scale-down, multipage, fixed-scale) per dare un
    contesto identificativo al foglio dati senza modificarne la struttura.

    Implementazione: legge tutti i valori e gli stili in memoria, cancella
    il foglio, scrive la riga titolo, poi riscrive i dati originali a partire
    dalla riga 2. Più verboso di `insert_rows(1)` ma evita i problemi noti
    di openpyxl con celle unite e settings di stampa quando si inseriscono
    righe in cima.

    Args:
        ws: worksheet openpyxl da modificare in-place.
        title: stringa del titolo da scrivere nella nuova riga 1.
        row_height_pt: altezza della riga titolo in punti (default 18).
    """
    from copy import copy as _copy
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.dimensions import RowDimension

    n_rows = ws.max_row
    n_cols = ws.max_column

    if n_rows < 1 or n_cols < 1:
        return

    # 1. Cattura valori, stili e altezze delle righe esistenti
    data = [
        [ws.cell(r, c).value for c in range(1, n_cols + 1)]
        for r in range(1, n_rows + 1)
    ]
    styles = [
        [
            (
                _copy(ws.cell(r, c).font),
                _copy(ws.cell(r, c).fill),
                _copy(ws.cell(r, c).alignment),
                _copy(ws.cell(r, c).border),
            )
            for c in range(1, n_cols + 1)
        ]
        for r in range(1, n_rows + 1)
    ]
    row_heights: dict[int, float] = {}
    for r in range(1, n_rows + 1):
        rd = ws.row_dimensions.get(r)
        if rd is not None and rd.height is not None:
            row_heights[r] = rd.height

    # 2. Reset del foglio
    ws.delete_rows(1, n_rows)
    ws.row_dimensions = type(ws.row_dimensions)(worksheet=ws)

    # 3. Riga titolo in posizione 1
    title_cell = ws.cell(1, 1, value=f"  {title}")
    title_cell.font = Font(size=12, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="305496")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    if n_cols > 1:
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=n_cols)
    ws.row_dimensions[1] = RowDimension(
        ws, ht=row_height_pt, customHeight=True,
    )

    # 4. Riscrivi i dati originali a partire dalla riga 2
    for r in range(n_rows):
        out_r = r + 2
        for c in range(n_cols):
            cell = ws.cell(out_r, c + 1, value=data[r][c])
            font, fill, alignment, border = styles[r][c]
            cell.font = font
            cell.fill = fill
            cell.alignment = alignment
            cell.border = border
        if (r + 1) in row_heights:
            ws.row_dimensions[out_r] = RowDimension(
                ws, ht=row_heights[r + 1], customHeight=True,
            )


def detect_row_blocks(ws) -> list[tuple[int, int]]:
    """Identifica blocchi di righe consecutive non vuote nel foglio,
    separati da una o più righe completamente vuote.

    Pensato per riconoscere strutture già presenti nei dati: es. più classi
    sullo stesso foglio separate da una riga vuota, sezioni multiple di una
    tabella, ecc.

    Returns:
        Lista di tuple `(start_row, end_row)` inclusive, 1-indexed.
        Esclude le righe vuote intermedie. Lista vuota se il foglio è vuoto.
    """
    n_rows = ws.max_row
    n_cols = ws.max_column

    if n_rows == 0 or n_cols == 0:
        return []

    def is_empty_row(r: int) -> bool:
        for c in range(1, n_cols + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            if isinstance(v, str) and v.strip() == "":
                continue
            return False
        return True

    blocks: list[tuple[int, int]] = []
    current_start: int | None = None

    for r in range(1, n_rows + 1):
        if is_empty_row(r):
            if current_start is not None:
                blocks.append((current_start, r - 1))
                current_start = None
        else:
            if current_start is None:
                current_start = r

    if current_start is not None:
        blocks.append((current_start, n_rows))

    return blocks


def apply_titles_to_row_blocks(
    ws,
    detected_blocks: list[tuple[int, int]],
    titles: list[str],
    title_row_height_pt: float = 18.0,
    separator_height_pt: float = 8.0,
) -> int:
    """Prepende una riga-titolo colorata all'inizio di ciascun blocco di righe
    già presente nel foglio (rilevato con detect_row_blocks).

    Es. se il foglio contiene 3 classi separate da righe vuote e l'utente
    passa --block "5°A" --block "5°B" --block "5°C", ognuna riceve la sua
    fascia colorata identificativa.

    Implementazione "delete + rewrite": legge tutto in memoria, pulisce il
    foglio, ricostruisce con titoli e separatori puliti. Le righe vuote
    originali tra blocchi vengono normalizzate a una sola riga di
    separatore stretto.

    Args:
        ws: worksheet openpyxl.
        detected_blocks: output di detect_row_blocks.
        titles: titoli da applicare in ordine. Se più di blocchi → extra
            ignorati; se meno → ultimi blocchi senza titolo.
        title_row_height_pt: altezza della riga-titolo.
        separator_height_pt: altezza della riga di separazione tra blocchi.

    Returns:
        Numero di blocchi a cui è stato effettivamente assegnato un titolo.
    """
    from copy import copy as _copy
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.dimensions import RowDimension

    n_rows = ws.max_row
    n_cols = ws.max_column

    if not detected_blocks or n_rows == 0 or n_cols == 0:
        return 0

    # 1. Cattura valori, stili e altezze del foglio originale
    data = [
        [ws.cell(r, c).value for c in range(1, n_cols + 1)]
        for r in range(1, n_rows + 1)
    ]
    styles = [
        [
            (
                _copy(ws.cell(r, c).font),
                _copy(ws.cell(r, c).fill),
                _copy(ws.cell(r, c).alignment),
                _copy(ws.cell(r, c).border),
            )
            for c in range(1, n_cols + 1)
        ]
        for r in range(1, n_rows + 1)
    ]
    row_heights: dict[int, float] = {}
    for r in range(1, n_rows + 1):
        rd = ws.row_dimensions.get(r)
        if rd is not None and rd.height is not None:
            row_heights[r] = rd.height

    # 2. Reset del foglio (celle, dimensioni, merge)
    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))
    ws.delete_rows(1, n_rows)
    ws.row_dimensions = type(ws.row_dimensions)(worksheet=ws)

    # 3. Riscrittura blocco per blocco
    out_row = 1
    titled = 0

    for block_idx, (start, end) in enumerate(detected_blocks):
        title = titles[block_idx] if block_idx < len(titles) else None

        # Separator tra blocchi consecutivi
        if block_idx > 0:
            ws.row_dimensions[out_row] = RowDimension(
                ws, ht=separator_height_pt, customHeight=True,
            )
            out_row += 1

        # Riga-titolo (se fornito per questo blocco)
        if title:
            title_cell = ws.cell(out_row, 1, value=f"  {title}")
            title_cell.font = Font(size=12, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill("solid", fgColor="305496")
            title_cell.alignment = Alignment(
                horizontal="left", vertical="center",
            )
            if n_cols > 1:
                ws.merge_cells(
                    start_row=out_row, start_column=1,
                    end_row=out_row, end_column=n_cols,
                )
            ws.row_dimensions[out_row] = RowDimension(
                ws, ht=title_row_height_pt, customHeight=True,
            )
            out_row += 1
            titled += 1

        # Righe-dati del blocco (preserva valori, stili, altezze)
        for src_r in range(start, end + 1):
            for c in range(1, n_cols + 1):
                cell = ws.cell(out_row, c, value=data[src_r - 1][c - 1])
                font, fill, alignment, border = styles[src_r - 1][c - 1]
                cell.font = font
                cell.fill = fill
                cell.alignment = alignment
                cell.border = border
            if src_r in row_heights:
                ws.row_dimensions[out_row] = RowDimension(
                    ws, ht=row_heights[src_r], customHeight=True,
                )
            out_row += 1

    return titled


def add_title_pages(wb, sheet_names: list[str], font_size: int = 28) -> list[str]:
    """Inserisce una pagina-copertina prima di ogni foglio del workbook.

    Ogni copertina è un foglio nuovo con una sola cella centrata
    contenente il nome del foglio originale, in font grande. Le impostazioni
    di stampa vengono copiate dal foglio dati corrispondente (stesso
    formato carta, stesso orientamento) così la copertina ha la stessa
    forma della pagina che la segue.

    Va chiamata DOPO che i fogli dati sono stati configurati con
    configure_print (per poter copiare paperSize e orientation reali).

    Args:
        wb: workbook openpyxl.
        sheet_names: nomi dei fogli dati originali (ordine in cui appariranno
            nel PDF). Le copertine vengono inserite davanti a ognuno.
        font_size: dimensione font del titolo in punti (default 60).

    Returns:
        Lista dei nomi dei fogli-copertina creati (utile per il debug).
    """
    from openpyxl.styles import Font, Alignment
    from openpyxl.worksheet.dimensions import RowDimension

    title_names: list[str] = []

    for sheet_name in sheet_names:
        data_ws = wb[sheet_name]
        # Posizione dove inserire: indice attuale del foglio dati.
        # create_sheet con index=N inserisce ALL'INIZIO della posizione N,
        # spostando il foglio dati di +1.
        target_idx = wb.sheetnames.index(sheet_name)

        # Nome univoco per non collidere con altri fogli.
        title_name = f"__title__{sheet_name}"[:31]   # Excel: max 31 char
        # Se per qualche motivo esiste già, aggiungo suffisso numerico.
        suffix = 1
        candidate = title_name
        while candidate in wb.sheetnames:
            candidate = f"{title_name}_{suffix}"[:31]
            suffix += 1
        title_name = candidate

        title_ws = wb.create_sheet(title_name, target_idx)
        title_names.append(title_name)

        # Contenuto: una sola cella con il nome del foglio in font grande
        cell = title_ws.cell(1, 1, value=sheet_name)
        cell.font = Font(size=font_size, bold=True, color="222222")
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Larghezza colonna e altezza riga ampie per dare risalto al titolo.
        # (la centratura della pagina avviene poi via print_options.)
        title_ws.column_dimensions["A"].width = 60
        title_ws.row_dimensions[1] = RowDimension(
            title_ws, ht=font_size * 1.5, customHeight=True,
        )

        # Stesso formato/orientamento del foglio dati che la segue.
        title_ws.page_setup.paperSize = data_ws.page_setup.paperSize
        title_ws.page_setup.orientation = data_ws.page_setup.orientation

        # Centra il titolo verticalmente e orizzontalmente nella pagina.
        title_ws.print_options.horizontalCentered = True
        title_ws.print_options.verticalCentered = True

        # Una sola pagina, no scaling automatico (la cella è già piccola
        # rispetto alla pagina; ce la facciamo stare con i margini).
        title_ws.page_setup.fitToWidth = 1
        title_ws.page_setup.fitToHeight = 1
        title_ws.sheet_properties.pageSetUpPr.fitToPage = True
        title_ws.page_setup.scale = None

        # Margini standard ampi: la copertina respira.
        title_ws.page_margins.left = 0.5
        title_ws.page_margins.right = 0.5
        title_ws.page_margins.top = 0.75
        title_ws.page_margins.bottom = 0.75

        # Print area minima.
        title_ws.print_area = "A1:A1"

    return title_names


def reset_print_state(ws):
    """Azzera tutto lo stato di stampa pre-esistente del worksheet.

    Senza questo passaggio, le impostazioni del file di origine possono
    sovrascrivere o sabotare quelle che imposteremo dopo. Cause tipiche
    di troncamento del PDF se non si fa il reset:

      - `print_area` impostata manualmente in Excel (Layout > Area di stampa)
      - `_xlnm.Print_Area` come named range (variante della precedente)
      - `page_setup.scale` valorizzato → ha precedenza sul fit-to-page
      - `col_breaks` / `row_breaks` manuali → limitano la stampa
      - `pageSetUpPr.autoPageBreaks = False` (modalità "anteprima interruzioni")
    """
    # 1. Print area su worksheet (attributo)
    ws.print_area = None

    # 2. Print area come named range a livello workbook
    #    (in alcune versioni di Excel/openpyxl le due cose coesistono)
    wb = ws.parent
    pa_name = f"'{ws.title}'!Print_Area"  # nome convenzionale
    if hasattr(wb, "defined_names"):
        # openpyxl >= 3.1 espone defined_names come dict-like
        names_to_remove = []
        for name in list(wb.defined_names):
            dn = wb.defined_names[name]
            # Heuristic: cerchiamo i Print_Area associati a questo foglio
            if name in ("_xlnm.Print_Area", "Print_Area"):
                # Vediamo se la destinazione è il foglio corrente
                try:
                    for sheet_title, _ in dn.destinations:
                        if sheet_title == ws.title:
                            names_to_remove.append(name)
                            break
                except Exception:
                    names_to_remove.append(name)
        for name in names_to_remove:
            del wb.defined_names[name]

    # 3. Scale fisso: se valorizzato, vince sul fit-to-page.
    #    None = "non impostato" (default behavior)
    ws.page_setup.scale = None

    # 4. Page break manuali: svuoto entrambe le liste interne (`.brk`).
    if ws.col_breaks is not None and hasattr(ws.col_breaks, "brk"):
        ws.col_breaks.brk = []
    if ws.row_breaks is not None and hasattr(ws.row_breaks, "brk"):
        ws.row_breaks.brk = []

    # 5. autoPageBreaks: se disabilitato (modalità anteprima interruzioni di
    #    Excel), LibreOffice non aggiunge interruzioni automatiche e tutto
    #    il contenuto resta "vincolato" alla prima pagina.
    ws.sheet_properties.pageSetUpPr.autoPageBreaks = True


def configure_print(
    ws,
    paper: str,
    orientation: str,
    margins: str = "narrow",
    repeat_header: bool = True,
    repeat_cols: int = 0,
    wide_strategy: str = "scale-down",
    fixed_scale: int = 85,
    wrap_block_starts: list[int] | None = None,
    auto_resolution_info: str | None = None,
    verbose: bool = False,
):
    """Imposta i parametri di stampa del worksheet.

    `paper`, `orientation` e `wide_strategy` devono già essere risolti
    (niente più "auto" qui: la risoluzione avviene a livello di pipeline
    in process_file, che ha gli elementi per decidere — paper, scala
    stimata, possibilità di wrap, ecc.).

    Strategie supportate:
      - "scale-down":  comprime tutto su 1 pagina di larghezza.
      - "multipage":   font pieno, colonne in eccesso su pagine successive.
      - "fixed-scale": scala fissa (fixed_scale %).
      - "wrap-rows":   il foglio è già stato ristrutturato da
                       wrap_sheet_into_rows; qui inseriamo solo i page
                       break passati in `wrap_block_starts`.
    """
    # Reset preventivo dello stato pre-esistente.
    reset_print_state(ws)

    # Formato carta
    if paper not in PAPER_SIZES:
        sys.exit(f"❌ Formato carta non valido: {paper}. "
                 f"Validi: {', '.join(PAPER_SIZES.keys())}")
    ws.page_setup.paperSize = PAPER_SIZES[paper]

    # Orientamento
    if orientation == "landscape":
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    else:
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

    # === Strategia di stampa ===
    pr = ws.sheet_properties.pageSetUpPr

    if wide_strategy == "scale-down":
        # Tutto su 1 pagina di larghezza, altezza libera (vedi nota su 9999).
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 9999
        ws.page_setup.scale = None
        pr.fitToPage = True
    elif wide_strategy == "multipage":
        # Font pieno: le colonne in eccesso scorrono su pagine successive.
        pr.fitToPage = False
        ws.page_setup.scale = 100
    elif wide_strategy == "fixed-scale":
        pr.fitToPage = False
        ws.page_setup.scale = fixed_scale
    elif wide_strategy == "wrap-rows":
        # Il foglio è già stato ristrutturato in blocchi. Inserisco un page
        # break PRIMA di ogni blocco che non entra nella pagina del precedente.
        #
        # Nota su openpyxl/OOXML: Break(id=N) inserisce il break DOPO la riga N
        # (la riga N resta nella pagina precedente). Per far iniziare il blocco
        # a riga `r` in una pagina nuova devo quindi usare Break(id=r-1).
        #
        # IMPORTANTE: niente fit-to-page qui. fitToPage=True combinato con i
        # page break manuali fa l'opposto di quello che serve: LibreOffice
        # ignora i break e scala tutto sulla singola pagina, ricreando
        # l'overflow illeggibile che il wrap voleva evitare.
        pr.fitToPage = False
        ws.page_setup.scale = 100
        if wrap_block_starts:
            from openpyxl.worksheet.pagebreak import Break
            ws.row_breaks.brk = [Break(id=r - 1) for r in wrap_block_starts]
    else:
        sys.exit(f"❌ Strategia non valida: {wide_strategy}")

    # Centra orizzontalmente
    ws.print_options.horizontalCentered = True

    # Margini (in pollici, standard openpyxl)
    if margins == "narrow":
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3
    elif margins == "normal":
        ws.page_margins.left = 0.7
        ws.page_margins.right = 0.7
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
    elif margins == "wide":
        ws.page_margins.left = 1.0
        ws.page_margins.right = 1.0
        ws.page_margins.top = 1.0
        ws.page_margins.bottom = 1.0

    # Ripeti riga di intestazione su ogni pagina (non utile con wrap-rows:
    # ogni blocco contiene già la propria intestazione fisica).
    if repeat_header and ws.max_row > 1 and wide_strategy != "wrap-rows":
        ws.print_title_rows = "1:1"

    # Per wrap-rows: stringi i margini verticali. Le righe vuote tra blocchi
    # forniscono già la separazione visiva interna; i margini larghi sopra/sotto
    # sprecano solo spazio utile per impacchettare più blocchi per pagina.
    # Il calcolo di blocks_per_page in wrap_sheet_into_rows assume questi
    # stessi margini (vedi page_height_for_wrap in process_file).
    if wide_strategy == "wrap-rows":
        ws.page_margins.top = 0.25
        ws.page_margins.bottom = 0.25

    # Ripeti le prime N colonne (es. ID/Cognome/Nome) su ogni pagina
    # orizzontale. Utile con multipage; con wrap-rows è ridondante perché
    # le colonne ancora sono già fisicamente ripetute in ogni blocco.
    if repeat_cols > 0 and wide_strategy != "wrap-rows":
        last_title_col = get_column_letter(repeat_cols)
        ws.print_title_cols = f"A:{last_title_col}"

    # Print area: usa calculate_dimension() per ottenere l'area dati REALE.
    try:
        dim = ws.calculate_dimension()  # es. "A1:AN101"
        if dim and dim != "A1:A1":
            ws.print_area = dim
    except Exception:
        if ws.max_row > 0 and ws.max_column > 0:
            last_cell = f"{get_column_letter(ws.max_column)}{ws.max_row}"
            ws.print_area = f"A1:{last_cell}"

    if verbose:
        strat_info = auto_resolution_info or wide_strategy
        print(f"     · strategia      = {strat_info}")
        print(f"     · print_area     = {ws.print_area}")
        if repeat_cols > 0:
            print(f"     · colonne ripet. = A:{get_column_letter(repeat_cols)}")
        if wrap_block_starts:
            print(f"     · page breaks    = righe {wrap_block_starts}")
        print(f"     · fitToPage      = {pr.fitToPage}")
        print(f"     · autoPageBreaks = {pr.autoPageBreaks}")
        print(f"     · scale          = {ws.page_setup.scale}")
        print(f"     · fitToWidth     = {ws.page_setup.fitToWidth}")
        print(f"     · fitToHeight    = {ws.page_setup.fitToHeight}")


# ─────────────────────────────────────────────────────────────────────────────
# Conversione LibreOffice
# ─────────────────────────────────────────────────────────────────────────────

def convert_with_libreoffice(
    input_path: Path,
    output_dir: Path,
    target_format: str,
    soffice: str,
    timeout: int = 180,
    verbose: bool = False,
) -> Path:
    """Converte un file con LibreOffice headless verso un formato target.

    Args:
        input_path: file di input (qualsiasi formato supportato da LibreOffice).
        output_dir: cartella di destinazione.
        target_format: nome del filtro di output, es. "pdf", "xlsx", "ods".
        soffice: percorso eseguibile soffice/libreoffice.
        timeout: timeout in secondi.
        verbose: stampa il comando eseguito.

    Returns:
        Path del file generato (stesso stem dell'input, nuova estensione).
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    # LibreOffice usa il proprio profilo utente; con più processi paralleli
    # può dare conflitti. Usiamo un profilo isolato e temporaneo.
    with tempfile.TemporaryDirectory(prefix="lo_profile_") as profile_dir:
        # `Path.as_uri()` produce un URI conforme su tutte le piattaforme:
        #   Linux:   /tmp/x        →  file:///tmp/x
        #   Windows: C:\Users\x    →  file:///C:/Users/x
        # Gestisce anche correttamente l'URL-encoding di spazi e accenti
        # (es. "Emanuele Rossi" → "Emanuele%20Rossi"). Senza questo, su
        # Windows LibreOffice rifiuta l'URI malformato segnalando
        # "bootstrap configuration file missing or corrupt".
        profile_uri = Path(profile_dir).as_uri()

        cmd = [
            soffice,
            "--headless",
            "--nologo",
            "--nofirststartwizard",
            "--nolockcheck",
            f"-env:UserInstallation={profile_uri}",
            "--convert-to", target_format,
            "--outdir", str(output_dir),
            str(input_path),
        ]

        if verbose:
            print(f"   $ {' '.join(cmd)}")

        try:
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=timeout,
                check=False,
            )
        except subprocess.TimeoutExpired:
            raise RuntimeError(
                f"Timeout LibreOffice ({timeout}s) su {input_path.name}"
            )

        if result.returncode != 0:
            raise RuntimeError(
                f"LibreOffice exit {result.returncode}\n"
                f"stderr: {result.stderr.strip()}"
            )

        if verbose and result.stdout.strip():
            print(f"   {result.stdout.strip()}")

    # Estensione attesa: primo token di target_format (es. "xlsx:Calc MS Excel 2007 XML")
    ext = target_format.split(":")[0].strip()
    expected = output_dir / f"{input_path.stem}.{ext}"
    if not expected.exists():
        raise RuntimeError(f"File non generato: atteso {expected}")
    return expected


def normalize_to_xlsx(
    input_path: Path,
    tmp_dir: Path,
    soffice: str,
    verbose: bool = False,
) -> Path:
    """Se l'input non è già xlsx/xlsm, lo converte in xlsx in tmp_dir.

    Returns:
        Path del file in formato xlsx/xlsm pronto per il preprocessing.
    """
    if input_path.suffix.lower() in NATIVE_EXTENSIONS:
        return input_path

    print(f"   🔁 Conversione preliminare {input_path.suffix} → .xlsx")
    return convert_with_libreoffice(
        input_path=input_path,
        output_dir=tmp_dir,
        target_format="xlsx",
        soffice=soffice,
        verbose=verbose,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Pipeline per file
# ─────────────────────────────────────────────────────────────────────────────

def process_file(
    input_path: Path,
    output_path: Path,
    soffice: str,
    paper: str,
    orientation: str,
    margins: str,
    max_col_width: float,
    min_col_width: float,
    repeat_header: bool,
    wide_strategy: str,
    repeat_cols: int,
    fixed_scale: int,
    min_scale: int,
    max_paper: str,
    wrap_safety: int,
    separator_rows: int,
    separator_height: float,
    block_titles_raw: list[str],
    block_skip_sheets: list[str],
    title_pages: bool,
    title_font_size: int,
    keep_temp: bool,
    verbose: bool,
) -> Path:
    """Pipeline completa di conversione di un singolo file."""
    print(f"\n📂 {input_path.name}")

    # Una sola tmp_dir per tutta la pipeline (normalizzazione + preprocessing).
    tmp_dir = Path(tempfile.mkdtemp(prefix="spreadsheet2pdf_"))

    try:
        # 0. Normalizza a xlsx se è un OpenDocument (.ods/.fods)
        xlsx_input = normalize_to_xlsx(
            input_path, tmp_dir, soffice, verbose=verbose,
        )

        # 1. Carica e prepara il workbook
        wb = load_workbook(xlsx_input, data_only=False)
        sheet_names = wb.sheetnames
        print(f"   📑 {len(sheet_names)} foglio/i: {', '.join(sheet_names)}")

        for name in sheet_names:
            ws = wb[name]
            if ws.max_row == 0 or ws.max_column == 0:
                print(f"   ⚠️  Foglio '{name}' vuoto, salto best-fit.")
                continue

            # Best-fit
            widths = compute_column_widths(
                ws,
                max_width=max_col_width,
                min_width=min_col_width,
            )
            total_w = sum(widths.values())

            # === Step 1: risolvi formato carta (auto con cap) ===
            final_paper = paper
            final_orient = orientation
            if paper == "auto" or orientation == "auto":
                ap, ao = auto_select_paper(ws.max_column, total_w, max_paper=max_paper)
                if paper == "auto":
                    final_paper = ap
                if orientation == "auto":
                    final_orient = ao

            # === Step 2: risolvi strategia ===
            effective_strategy = wide_strategy
            est_scale = None
            if wide_strategy == "auto":
                est_scale = estimate_fit_scale(
                    total_w, ws.max_column, final_paper, final_orient, margins,
                )
                # Sopra min_scale comprimo; sotto wrappo (più leggibile a video).
                effective_strategy = (
                    "scale-down" if est_scale >= min_scale else "wrap-rows"
                )

            auto_info = None
            if wide_strategy == "auto":
                auto_info = (f"auto→{effective_strategy} "
                             f"(scala stimata {est_scale}% su {final_paper.upper()})")
            # Marker [+ --block] solo se i titoli verranno effettivamente
            # applicati a questo foglio (i fogli in --block-skip non lo sono).
            skip_check = any(s.lower() == name.lower() for s in block_skip_sheets)
            if block_titles_raw and not skip_check:
                auto_info = (auto_info + " [+ --block]" if auto_info
                             else f"{effective_strategy} [+ --block]")

            # === Step 2b: --block applica titoli ai blocchi presenti nei dati ===
            # Distinguo due casi:
            # 1) I dati hanno PIÙ blocchi di righe (separati da righe vuote):
            #    es. più classi sullo stesso foglio. Allora --block titola
            #    ciascuno di questi blocchi, indipendentemente dalla strategia.
            # 2) I dati hanno un solo blocco unico:
            #    - in wrap-rows i titoli vanno ai blocchi-COLONNA del wrap
            #      (gestiti da wrap_sheet_into_rows più sotto);
            #    - in scale-down/multipage/fixed-scale il primo titolo
            #      diventa una riga-header globale del foglio.
            title_prepended = False
            data_blocks_titled = False

            # Lista esclusioni: confronto case-insensitive del nome foglio.
            skip_this_sheet = any(
                s.lower() == name.lower() for s in block_skip_sheets
            )

            if block_titles_raw and skip_this_sheet:
                if verbose:
                    print(f"   ▸ '{name}': escluso dalla logica --block "
                          f"(--block-skip)")
            elif block_titles_raw:
                detected = detect_row_blocks(ws)
                if len(detected) > 1:
                    # Caso 1: blocchi multipli nei dati
                    titled = apply_titles_to_row_blocks(
                        ws, detected, block_titles_raw,
                        separator_height_pt=separator_height,
                    )
                    print(f"   ↳ '{name}': rilevati {len(detected)} blocchi "
                          f"nei dati, {titled} titolati")
                    if len(block_titles_raw) > len(detected):
                        extra = block_titles_raw[len(detected):]
                        print(f"   ⚠️  titoli extra ignorati: {extra}")
                    elif len(block_titles_raw) < len(detected):
                        n_missing = len(detected) - len(block_titles_raw)
                        print(f"   ⚠️  {n_missing} blocchi senza titolo "
                              f"(--block forniti: {len(block_titles_raw)})")
                    # wrap-rows non è compatibile con blocchi di righe già
                    # presenti nei dati (assumerebbe foglio tabulare unico):
                    # ripiego su scale-down per non rompere la struttura.
                    if effective_strategy == "wrap-rows":
                        print(f"   ⚠️  '{name}': wrap-rows incompatibile con "
                              f"blocchi multipli nei dati, uso scale-down")
                        effective_strategy = "scale-down"
                        # Aggiorno il marker diagnostico per coerenza con
                        # la strategia effettivamente applicata.
                        auto_info = f"{effective_strategy} [+ --block, dati multi-blocco]"
                    data_blocks_titled = True
                elif effective_strategy != "wrap-rows":
                    # Caso 2: foglio a blocco unico + strategia non-wrap →
                    # primo titolo come header globale.
                    if len(block_titles_raw) > 1:
                        print(f"   ⚠️  '{name}': {len(block_titles_raw)} titoli "
                              f"forniti ma 1 solo blocco nei dati; "
                              f"uso '{block_titles_raw[0]}', ignoro "
                              f"{block_titles_raw[1:]}")
                    prepend_header_row(ws, block_titles_raw[0])
                    title_prepended = True
                # else: 1 solo blocco + wrap-rows → wrap applicherà i titoli
                # per blocco-colonne (gestito dopo)

            # === Step 3: se wrap-rows, ristruttura il foglio ===
            wrap_block_starts = None
            if effective_strategy == "wrap-rows":
                if repeat_cols == 0:
                    print(f"   ⚠️  '{name}': wrap-rows senza --repeat-cols → "
                          f"i blocchi non avranno colonne identificative")

                printable = printable_width_mm(final_paper, final_orient, margins)
                anchor_w_mm = sum(
                    char_to_mm(widths.get(c, 8))
                    for c in range(1, repeat_cols + 1)
                )
                available = printable - anchor_w_mm

                if available <= 0:
                    print(f"   ⚠️  '{name}': le {repeat_cols} colonne ancora "
                          f"occupano più della pagina; ripiego su scale-down")
                    effective_strategy = "scale-down"
                else:
                    # Altezza utile della pagina per wrap-rows: usiamo margini
                    # verticali stretti (12.7mm = 0.25"+0.25") che configure_print
                    # applicherà al worksheet. Va sincronizzato qui per stimare
                    # correttamente quanti blocchi entrano per pagina.
                    short_mm, long_mm = PAPER_DIMENSIONS_MM.get(
                        final_paper, PAPER_DIMENSIONS_MM["a4"]
                    )
                    page_h = (short_mm if final_orient == "landscape" else long_mm) - 12.7

                    wrap_block_starts, n_blocks = wrap_sheet_into_rows(
                        ws,
                        anchor_cols=repeat_cols,
                        available_data_width_mm=available,
                        col_widths_chars=widths,
                        safety_margin=wrap_safety / 100.0,
                        page_height_mm=page_h,
                        separator_rows=separator_rows,
                        separator_height_pt=separator_height,
                        block_titles=block_titles_raw or None,
                    )
                    if n_blocks > 1:
                        n_pages = len(wrap_block_starts) + 1
                        print(f"   ↳ '{name}': wrap su {n_blocks} blocchi "
                              f"distribuiti su {n_pages} pagine "
                              f"({ws.max_column} colonne nel layout finale)")

            # === Step 4: configure print con strategia già risolta ===
            configure_print(
                ws,
                paper=final_paper,
                orientation=final_orient,
                margins=margins,
                repeat_header=repeat_header,
                repeat_cols=repeat_cols,
                wide_strategy=effective_strategy,
                fixed_scale=fixed_scale,
                wrap_block_starts=wrap_block_starts,
                auto_resolution_info=auto_info,
                verbose=verbose,
            )

            # Aggiorno print_title_rows in base a cosa è stato applicato:
            # - title_prepended: ho aggiunto UNA riga-titolo in cima, l'header
            #   colonne è scivolato in riga 2 → ripeto "2:2".
            # - data_blocks_titled: ci sono MULTIPLE intestazioni colonne nel
            #   foglio (una per blocco-dati), nessuna ripetizione globale
            #   avrebbe senso → None.
            if title_prepended and repeat_header:
                ws.print_title_rows = "2:2"
            elif data_blocks_titled:
                ws.print_title_rows = None

            if verbose:
                print(f"   ▸ '{name}': {ws.max_column} col × {ws.max_row} righe "
                      f"→ {final_paper.upper()} {final_orient}")

        # 1.5 Pagine titolo: inserisco una copertina prima di ogni foglio,
        # solo se richiesto e se c'è più di un foglio.
        if title_pages and len(sheet_names) > 1:
            created = add_title_pages(wb, sheet_names, font_size=title_font_size)
            print(f"   📑 Aggiunte {len(created)} pagine-copertina "
                  f"(disabilita con --no-title-pages)")

        # 2. Salva la copia preparata in tmp.
        # Nome distinto per non collidere con un eventuale xlsx normalizzato.
        prepared = tmp_dir / f"{input_path.stem}_prepared.xlsx"
        wb.save(prepared)
        wb.close()

        # 3. Conversione LibreOffice
        print(f"   🔄 Conversione in PDF...")
        generated_pdf = convert_with_libreoffice(
            input_path=prepared,
            output_dir=output_path.parent,
            target_format="pdf",
            soffice=soffice,
            verbose=verbose,
        )

        # Rinomina se l'output richiesto è diverso dal nome predefinito
        if generated_pdf.name != output_path.name:
            generated_pdf.replace(output_path)
            generated_pdf = output_path

        size_kb = generated_pdf.stat().st_size / 1024
        print(f"   ✅ {generated_pdf.name} ({size_kb:.1f} KB)")
        return generated_pdf

    finally:
        if keep_temp:
            print(f"   🗂️  File temporanei conservati in: {tmp_dir}")
        else:
            shutil.rmtree(tmp_dir, ignore_errors=True)


# ─────────────────────────────────────────────────────────────────────────────
# Main / CLI
# ─────────────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(
        prog="spreadsheet-to-pdf",
        description="Converte fogli elettronici (Excel e OpenDocument, "
                    "multifoglio) in PDF con auto-fit colonne.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Esempi:\n"
            "  spreadsheet-to-pdf.py report.xlsx\n"
            "  spreadsheet-to-pdf.py registro.ods\n"
            "  spreadsheet-to-pdf.py dati.xlsx -o out.pdf --paper a3\n"
            "  spreadsheet-to-pdf.py *.xlsx *.ods --outdir ./pdf/ -v\n"
        ),
    )
    p.add_argument(
        "inputs",
        nargs="+",
        help="File da convertire: .xlsx, .xlsm, .ods, .fods "
             "(uno o più, supporta glob della shell).",
    )
    p.add_argument(
        "-o", "--output",
        help="PDF di output (solo se input è un file singolo). "
             "Default: stesso nome con estensione .pdf nella stessa cartella.",
    )
    p.add_argument(
        "--outdir",
        help="Cartella di output (per batch). Default: stessa cartella del sorgente.",
    )
    p.add_argument(
        "--paper",
        default="auto",
        choices=["auto"] + list(PAPER_SIZES.keys()),
        help="Formato carta (default: auto, basato sul numero di colonne).",
    )
    p.add_argument(
        "--orientation",
        default="auto",
        choices=["auto", "portrait", "landscape"],
        help="Orientamento (default: auto, di solito landscape).",
    )
    p.add_argument(
        "--margins",
        default="narrow",
        choices=["narrow", "normal", "wide"],
        help="Margini di stampa (default: narrow).",
    )
    p.add_argument(
        "--max-col-width",
        type=float,
        default=DEFAULT_MAX_COL_WIDTH,
        help=f"Larghezza massima colonna in caratteri (default: {DEFAULT_MAX_COL_WIDTH}). "
             "Utile per evitare colonne enormi su testi lunghi.",
    )
    p.add_argument(
        "--min-col-width",
        type=float,
        default=DEFAULT_MIN_COL_WIDTH,
        help=f"Larghezza minima colonna (default: {DEFAULT_MIN_COL_WIDTH}).",
    )
    p.add_argument(
        "--no-repeat-header",
        action="store_true",
        help="Non ripetere la prima riga come intestazione su ogni pagina.",
    )
    p.add_argument(
        "--wide-strategy",
        default="auto",
        choices=["auto", "scale-down", "multipage", "fixed-scale", "wrap-rows"],
        help="Comportamento per fogli più larghi della pagina (default: auto):\n"
             "  auto        sotto min_scale ripiega su wrap-rows, sopra scale-down\n"
             "  scale-down  comprime tutto su 1 pagina di larghezza (testo piccolo)\n"
             "  multipage   font pieno, colonne in eccesso su pagine successive\n"
             "  fixed-scale scala fissa (--fixed-scale), impaginazione naturale\n"
             "  wrap-rows   sposta colonne in eccesso in righe successive,\n"
             "              ripetendo le prime --repeat-cols colonne (ottima\n"
             "              per consultazione a video)",
    )
    p.add_argument(
        "--max-paper",
        default="a3",
        choices=["a4", "a3", "a2", "a1"],
        help="Cap del formato carta in modalità auto (default: a3). "
             "Oltre questo limite la strategia auto preferisce wrap-rows "
             "anziché crescere ad A2/A1 illeggibili a schermo.",
    )
    p.add_argument(
        "--repeat-cols",
        type=int,
        default=0,
        metavar="N",
        help="Ripete le prime N colonne (es. ID/Cognome/Nome) come anchor "
             "in ogni pagina/blocco. Consigliato con multipage e wrap-rows "
             "(default: 0 = nessuna).",
    )
    p.add_argument(
        "--fixed-scale",
        type=int,
        default=85,
        metavar="PCT",
        help="Scala percentuale per --wide-strategy fixed-scale (default: 85).",
    )
    p.add_argument(
        "--min-scale",
        type=int,
        default=95,
        metavar="PCT",
        help="Soglia sotto cui 'auto' passa da scale-down a wrap-rows "
             "(default: 95). Più alta = più probabile wrap, più leggibile.",
    )
    p.add_argument(
        "--wrap-safety",
        type=int,
        default=12,
        metavar="PCT",
        help="Margine di sicurezza percentuale per il wrap (default: 12). "
             "Riduce la larghezza utile per blocco per coprire l'errore "
             "della stima font→mm; alzalo se i blocchi finiscono ancora "
             "su pagine multiple orizzontali.",
    )
    p.add_argument(
        "--separator-rows",
        type=int,
        default=1,
        metavar="N",
        help="Righe vuote di separazione tra blocchi consecutivi nella stessa "
             "pagina con wrap-rows (default: 1).",
    )
    p.add_argument(
        "--separator-height",
        type=float,
        default=8.0,
        metavar="PT",
        help="Altezza in punti delle righe-separatore con wrap-rows "
             "(default: 8.0, default Excel è ~15). Più basso = più blocchi "
             "per pagina.",
    )
    p.add_argument(
        "--soffice",
        help="Percorso esplicito all'eseguibile soffice/libreoffice.",
    )
    p.add_argument(
        "--block",
        action="append",
        default=[],
        metavar="TITOLO",
        help="Aggiunge un titolo colorato a un blocco di righe del foglio. "
             "Ripetibile. Lo script rileva automaticamente i blocchi nei "
             "dati (sezioni di righe separate da righe vuote) e applica i "
             "titoli in ordine, indipendentemente dalla --wide-strategy. "
             "Se ne rileva uno solo, in wrap-rows i titoli vengono assegnati "
             "ai blocchi-colonne generati dal wrap; nelle altre strategie "
             "il primo titolo diventa un header globale del foglio. "
             "Esempio: --block '5°A' --block '5°B' --block '5°C'.",
    )
    p.add_argument(
        "--block-skip",
        action="append",
        default=[],
        metavar="FOGLIO",
        help="Nome di un foglio (case-insensitive) da escludere dalla logica "
             "--block. Ripetibile. Utile per workbook che mischiano fogli "
             "strutturati a blocchi (a cui applicare i titoli) con altri "
             "fogli di riassunto/note che non li hanno. I fogli esclusi "
             "vengono comunque convertiti, ma senza titoli/header.",
    )
    p.add_argument(
        "--no-title-pages",
        action="store_true",
        help="Disabilita le pagine-copertina con il nome del foglio "
             "inserite prima di ogni foglio quando il workbook ne contiene "
             "più di uno (di default sono attive).",
    )
    p.add_argument(
        "--title-font-size",
        type=int,
        default=28,
        metavar="PT",
        help="Dimensione del titolo nelle pagine-copertina (default: 28 pt).",
    )
    p.add_argument(
        "--keep-temp",
        action="store_true",
        help="Conserva il file XLSX intermedio (debug).",
    )
    p.add_argument(
        "-v", "--verbose",
        action="store_true",
        help="Output dettagliato.",
    )
    return p.parse_args()


def main():
    args = parse_args()

    # Raccogli i file in input
    inputs: list[Path] = []
    for pattern in args.inputs:
        path = Path(pattern)
        if path.is_file():
            inputs.append(path)
        else:
            # Supporto glob anche se la shell non l'ha espanso
            matches = list(Path().glob(pattern))
            if matches:
                inputs.extend(p for p in matches if p.is_file())
            else:
                print(f"⚠️  Nessun file trovato: {pattern}", file=sys.stderr)

    if not inputs:
        sys.exit("❌ Nessun file XLSX valido in input.")

    # Filtra solo formati supportati
    valid_inputs = [p for p in inputs
                    if p.suffix.lower() in SUPPORTED_EXTENSIONS]
    skipped = set(inputs) - set(valid_inputs)
    for s in skipped:
        print(f"⚠️  Estensione non supportata, salto: {s.name}", file=sys.stderr)

    if not valid_inputs:
        sys.exit("❌ Nessun file supportato in input "
                 f"(estensioni accettate: {', '.join(sorted(SUPPORTED_EXTENSIONS))}).")

    # Validazione -o vs --outdir
    if args.output and len(valid_inputs) > 1:
        sys.exit("❌ --output può essere usato solo con un file in input. "
                 "Usa --outdir per batch.")

    # Risolvi soffice
    soffice = find_soffice(args.soffice)
    print(f"🔧 LibreOffice: {soffice}")

    # Header
    print("=" * 60)
    print(f"  📊 Foglio elettronico → PDF  ({len(valid_inputs)} file)")
    print("=" * 60)

    successes, failures = 0, 0
    for input_path in valid_inputs:
        # Determina percorso output
        if args.output:
            output_path = Path(args.output).resolve()
        else:
            outdir = Path(args.outdir).resolve() if args.outdir else input_path.parent
            output_path = outdir / (input_path.stem + ".pdf")

        try:
            process_file(
                input_path=input_path.resolve(),
                output_path=output_path,
                soffice=soffice,
                paper=args.paper,
                orientation=args.orientation,
                margins=args.margins,
                max_col_width=args.max_col_width,
                min_col_width=args.min_col_width,
                repeat_header=not args.no_repeat_header,
                wide_strategy=args.wide_strategy,
                repeat_cols=args.repeat_cols,
                fixed_scale=args.fixed_scale,
                min_scale=args.min_scale,
                max_paper=args.max_paper,
                wrap_safety=args.wrap_safety,
                separator_rows=args.separator_rows,
                separator_height=args.separator_height,
                block_titles_raw=args.block,
                block_skip_sheets=args.block_skip,
                title_pages=not args.no_title_pages,
                title_font_size=args.title_font_size,
                keep_temp=args.keep_temp,
                verbose=args.verbose,
            )
            successes += 1
        except Exception as e:
            failures += 1
            print(f"   ❌ Errore: {e}", file=sys.stderr)
            if args.verbose:
                import traceback
                traceback.print_exc()

    print("\n" + "=" * 60)
    print(f"  ✨ Completato: {successes} OK, {failures} errori")
    print("=" * 60)
    sys.exit(0 if failures == 0 else 1)


if __name__ == "__main__":
    main()